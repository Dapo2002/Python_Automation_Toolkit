# Iterates through Excel files in a directory and converts each worksheet into an individual CSV file.
# Uses openpyxl to parse cell data and the csv module to handle formatting and file output.

import os
import openpyxl
import csv
from openpyxl.utils import get_column_letter

os.makedirs(r".\spreadsheetsToCsvFolders", exist_ok=True)
excelPath = r".\excelSpreadsheets"
csvDestinationPath = r".\spreadsheetsToCsvFolders"

for excelFile in os.listdir(excelPath):

    # Skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(os.path.join(excelPath, excelFile))

    # Loop through every sheet in the workbook.
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]

        # Create the CSV filename from the Excel filename and sheet title.
        csvFile = open(os.path.join(csvDestinationPath, f"{excelFile.removesuffix('.xlsx')}_{sheetName}.csv"),
                       'w', newline='')

        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list

            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):

                # Append each cell's data to rowData.
                cellValue = sheet[f'{get_column_letter(colNum)}{rowNum}']
                rowData.append(cellValue.value)
            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFile.close()
