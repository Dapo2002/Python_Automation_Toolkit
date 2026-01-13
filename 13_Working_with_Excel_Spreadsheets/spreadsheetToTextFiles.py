# A reverse-etl utility that decomposes a spreadsheet into multiple text 
# documents, mapping each column's contents to a separate file within a target 
# directory.

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

wb = openpyxl.load_workbook('textFilesToSpreadsheet.xlsx')
sheet = wb.active
p = Path(r"C:\Users\User\OneDrive\User\PycharmProjects"
         r"\workingWithExcelSpreadsheets\Spreadsheet to Text Files example")
list_outer = []
for col in range(1, sheet.max_column + 1):
    list_inner = []
    for row in range(1, sheet.max_row + 1):
        rowValues = sheet[f'{get_column_letter(col)}{row}'].value
        list_inner.append(rowValues)
    list_outer.append(list_inner)

for i in range(len(list_outer)):
    file = open(p / f'file_{i + 1}.txt', 'w')
    for j in range(len(list_outer[i])):
        file.write(f'{list_outer[i][j]}')
    file.close()
