# A high-performance multiplication table generator that populates an NxN grid 
# with dynamic Excel formulas, utilizing pane freezing and bold formatting for 
# header clarity.

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import sys

wb = openpyxl.Workbook()
sheet = wb.active
N = int(sys.argv[1])
boldFont = Font(bold=True)
sheet.freeze_panes = 'B2'
for i in range(1, N+1):
    sheet['' + get_column_letter(i+1) + '1'].font = boldFont
    sheet['A' + str(i+1)].font = boldFont
    sheet['' + get_column_letter(i+1) + '1'] = i
    sheet['A' + str(i+1)] = i

for i in sheet[f'b2:{get_column_letter(N+1)}{N+1}']:
    for j in i:
        sheet[j.coordinate] = f'=A{j.row} * {j.column_letter}1'

wb.save('multiplicationTableSheet.xlsx')
