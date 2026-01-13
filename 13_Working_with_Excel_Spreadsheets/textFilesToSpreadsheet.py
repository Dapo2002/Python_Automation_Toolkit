#  textFilesToSpreadsheet.py
#  It reads every `.txt` file in a folder and writes each file’s lines
#  into its own column of a new Excel spreadsheet.


from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

p = Path(r'C:\Users\User\OneDrive\User\PycharmProjects\workingWithExcelSpreadsheets'
         r'\Text Files to Spreadsheet Files example')
wb = openpyxl.Workbook()
sheet = wb.active
column = 1
for file in p.glob('*.txt'):
    fileContent = open(file)
    fileLines = fileContent.readlines()
    fileContent.close()

    for row in range(len(fileLines)):
        sheet[f'{get_column_letter(column)}{row + 1}'] = fileLines[row]
    column += 1

wb.save('textFilesToSpreadsheet.xlsx')
