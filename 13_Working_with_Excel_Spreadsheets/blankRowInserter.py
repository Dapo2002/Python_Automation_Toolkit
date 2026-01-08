# A spreadsheet utility that performs row-insertion by migrating data from a 
# source file to a new workbook, introducing an M-row offset at index N.

import sys
import openpyxl
from openpyxl.utils import get_column_letter


wb = openpyxl.load_workbook(f'{sys.argv[3]}')
sheet = wb.active
wb1 = openpyxl.Workbook()
sheet1 = wb1.active

writePointer = 1
N = int(sys.argv[1])
M = int(sys.argv[2])

for i in range(1, sheet.max_row + 1):
    if i == N:
        writePointer += M
    for j in range(1, sheet.max_column + 1):
        cellValue3 = sheet[f'{get_column_letter(j)}{i}'].value
        sheet1[f'{get_column_letter(j)}{writePointer}'] = cellValue3
    writePointer += 1

wb1.save(f'{sys.argv[3].replace(".xlsx", "")}Inserted.xlsx')
