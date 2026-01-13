# A spreadsheet utility that performs a matrix transposition, inverting the 
# coordinate system so that row-based records are converted into column-based 
# fields.

import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet1']
wb1 = openpyxl.Workbook()
sheet1 = wb1.active

sheetData = []

for row in range(1, sheet.max_row + 1):
    rowList = []
    for col in range(1, sheet.max_column + 1):
        cellValue = sheet[f'{get_column_letter(col)}{row}'].value
        rowList.append(cellValue)
    sheetData.append(rowList)

for y in range(len(sheetData[0])):
    for x in range(len(sheetData)):
        cellValue = sheetData[x][y]
        sheet1[f'{get_column_letter(x + 1)}{y + 1}'] = cellValue

wb1.save('produceSales_Inverted.xlsx')
